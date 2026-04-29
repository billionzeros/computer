#!/usr/bin/env python3
"""Recalculate every formula in an .xlsx file via headless LibreOffice and
report any cells that ended up as Excel error values.

openpyxl writes formula strings but does not evaluate them. Most viewers
read cached values, so a freshly-written workbook shows blank formula
cells until something opens it and saves it again. This script does that
round-trip and then walks every cell looking for #REF!, #DIV/0!, #VALUE!,
#NAME?, #N/A, #NUM!, #NULL!.

Usage:
  python recalc.py path/to/file.xlsx [timeout_seconds]

Output (stdout, JSON):
  {
    "status": "success" | "errors_found" | "failed",
    "total_formulas": <int>,
    "total_errors": <int>,
    "error_summary": { "#REF!": { "count": N, "locations": ["Sheet1!B5", ...] }, ... },
    "message": <only when status == "failed">
  }
"""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import tempfile

EXCEL_ERROR_PREFIXES = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!")


def find_soffice() -> str:
    candidate = shutil.which("soffice")
    if candidate:
        return candidate
    for fallback in (
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
        "/snap/bin/libreoffice",
    ):
        if os.path.exists(fallback):
            return fallback
    raise RuntimeError(
        "LibreOffice not found. Install it (macOS: `brew install --cask libreoffice`; "
        "Linux: `apt install libreoffice`) and ensure `soffice` is on PATH."
    )


def soffice_recalc(input_path: str, output_path: str, timeout: int) -> None:
    """Open input_path in headless LibreOffice Calc and write a recalculated copy."""
    binary = find_soffice()
    with tempfile.TemporaryDirectory(prefix="anton-soffice-") as profile_dir:
        outdir = os.path.dirname(output_path) or "."
        os.makedirs(outdir, exist_ok=True)
        cmd = [
            binary,
            "--headless",
            "--calc",
            f"-env:UserInstallation=file://{profile_dir}",
            "--convert-to",
            "xlsx",
            "--outdir",
            outdir,
            input_path,
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice exited {result.returncode}: {result.stderr.strip() or result.stdout.strip()}"
            )
        produced = os.path.join(outdir, os.path.basename(input_path))
        if produced != output_path:
            shutil.move(produced, output_path)


def scan_errors(path: str) -> tuple[int, int, dict]:
    """Walk every cell, count formulas and Excel errors. Returns (formulas, errors, summary)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError(
            "openpyxl is not installed. Install it with: pip install openpyxl"
        )
    wb = load_workbook(path, data_only=False)
    cached = load_workbook(path, data_only=True)
    formula_count = 0
    summary: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cached_ws = cached[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    cached_value = cached_ws[cell.coordinate].value
                    if isinstance(cached_value, str) and any(
                        cached_value.startswith(p) for p in EXCEL_ERROR_PREFIXES
                    ):
                        bucket = summary.setdefault(
                            cached_value, {"count": 0, "locations": []}
                        )
                        bucket["count"] += 1
                        if len(bucket["locations"]) < 25:
                            bucket["locations"].append(f"{sheet_name}!{cell.coordinate}")

    error_count = sum(b["count"] for b in summary.values())
    return formula_count, error_count, summary


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(json.dumps({"status": "failed", "message": "usage: recalc.py <file.xlsx> [timeout]"}))
        return 2

    input_path = argv[1]
    timeout = int(argv[2]) if len(argv) > 2 else 90

    if not os.path.exists(input_path):
        print(json.dumps({"status": "failed", "message": f"file not found: {input_path}"}))
        return 2

    try:
        with tempfile.TemporaryDirectory(prefix="anton-recalc-") as work:
            tmp = os.path.join(work, os.path.basename(input_path))
            soffice_recalc(input_path, tmp, timeout)
            shutil.move(tmp, input_path)
        formulas, errors, summary = scan_errors(input_path)
    except Exception as e:
        print(json.dumps({"status": "failed", "message": str(e)}))
        return 1

    payload = {
        "status": "errors_found" if errors > 0 else "success",
        "total_formulas": formulas,
        "total_errors": errors,
    }
    if summary:
        payload["error_summary"] = summary
    print(json.dumps(payload, indent=2))
    return 0 if errors == 0 else 3


if __name__ == "__main__":
    sys.exit(main(sys.argv))
